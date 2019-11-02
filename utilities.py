#! /usr/bin/env python3
# coding=utf-8

import os
import subprocess
from collections import namedtuple
import logging
import io


from lxml import etree as ET
import openpyxl


def parse_xlsx_file(xlsx_file):
    try:
        workbook = openpyxl.load_workbook(xlsx_file)
    except openpyxl.utils.exceptions.InvalidFileException:
        logging.fatal(f"'{xlsx_file}' does not appear to be a valid xlsx Excel file. \n Program cancelled")
        quit()
    mappings = parse_mappings(workbook)
    metadata = parse_metadata(workbook)
    xsls = parse_xsls(workbook)
    return mappings, metadata, xsls


def parse_mappings(workbook):
    try:
        mappings_sheet = workbook.get_sheet_by_name('Mappings')
    except KeyError:
        logging.fatal(f"""Could not find worksheet "Mappings" in the xlsx file. \n Program cancelled""")
        quit()
    mappings = {shorten(row[0].value): row[1].value for row in mappings_sheet.iter_rows()}
    return mappings


def parse_metadata(workbook):
    try:
        sheet = workbook.get_sheet_by_name('Metadata')
    except KeyError:
        logging.fatal(f"""Could not find worksheet "Metadata" in the xlsx file. \n Program cancelled""")
        quit()
    max_columns = count_active_columns(sheet)
    metadata = dict()
    for row_num, row in enumerate(sheet.iter_rows(max_col=max_columns)):
        if row_num == 0:
            headers = [shorten(i.value) for i in row]
            continue
        values = (i.value for i in row)
        item = dict(zip(headers, values))
        item["Row"] = row_num + 1
        metadata[row_num + 1] = item  # 1-indexing so that key matches spreadsheet row number
    return metadata


def parse_xsls(workbook):
    try:
        sheet = workbook.get_sheet_by_name('Xsls')
    except KeyError:
        logging.fatal(f"""Could not find worksheet "Xsls" in the xlsx file. \n Program cancelled""")
        quit()
    max_columns = count_active_columns(sheet)
    xsls = [i[0].value for i in sheet.iter_rows(max_col=max_columns) if i[0].value]
    return xsls


def shorten(fullname):
    return ''.join([i for i in fullname if i.isalnum()])


def count_active_columns(worksheet):
    row_1 = [i for i in worksheet.iter_rows(min_row=1, max_row=1)][0]
    return len({cell.value for cell in row_1}) - 1


def fix_permissions():
    all_files = [
        os.path.join(root, file)
        for root, dirs, files in os.walk('../cDM_to_mods/')
        for file in files
    ]
    all_dirs = [
        os.path.join(root, dir)
        for root, dirs, files in os.walk('../cDM_to_mods')
        for dir in dirs
    ]
    for file in all_files:
        subprocess.run(['chmod', '664', file])
    for dir in all_dirs:
        subprocess.run(['chmod', '775', dir])


def setup_logging():
    formatter = logging.Formatter('%(name)-12s: %(levelname)-8s %(message)s')
    logging.basicConfig(filename='log.txt',
                        level=logging.INFO,
                        format='%(asctime)s: %(levelname)-8s %(message)s',
                        datefmt='%m/%d/%Y %I:%M:%S %p')
    console = logging.StreamHandler()
    console.setLevel(logging.INFO)
    console.setFormatter(formatter)
    logging.getLogger('').addHandler(console)
    logging_string = io.StringIO()
    string_handler = logging.StreamHandler(logging_string)
    string_handler.setLevel(logging.DEBUG)
    string_handler.setFormatter(formatter)
    logging.getLogger('').addHandler(string_handler)
    return logging_string


def group_by_simple_cpd(metadata):
    simples, compounds = list(), dict()
    child_of = False
    for row_num, item_metadata in sorted(metadata.items()):

        # Logic of this function --
        # if this row has nothing in the Child cell
        #    & there is a next row
        #    & that next row has info in the Child cell,
        #    then
        #       this row is a parent
        #       set child_of flag to the parent identifier
        # else if this row has info in the Child cell
        #    then
        #       it is a child object
        #       its parent's name is in the child_of flag
        # Otherwise
        #       this item is a simple object
        #       set child_of flag to False.

        if (
            not item_metadata['Child'] and
            metadata.get(row_num + 1) and
            metadata.get(row_num + 1)['Child']
        ):
            child_of = item_metadata['Identifier']
            # catch fire if overlapping parent ids
            if compounds.get(item_metadata['Identifier']):
                logging.fatal(f"two parents in spreadsheet with id: {item_metadata['Identifier']} \n Program cancelled")
                quit()
            compounds[item_metadata['Identifier']] = {'parent': item_metadata, }
        elif item_metadata['Child']:
            # catch fire if two children with same id
            if compounds[child_of].get(int(item_metadata['Child'])):
                logging.fatal(f"two children in spreadsheet with id: {child_of} {item_metadata['Child']} \n Program cancelled")
                quit()
            item_metadata['Parent'] = child_of
            compounds[child_of][int(item_metadata['Child'])] = item_metadata
        else:
            simples.append(item_metadata)
    return simples, compounds


class MonographTitleCombiner:
    def __init__(self, alias_data_dir):
        self.alias_data_dir = alias_data_dir
        self.monograph_pointer_newtitle = dict()
        self.current_stucture_file = None
        self.main()

    def main(self):
        structure_files = [os.path.join(root, file)
                           for root, dirs, files in os.walk(self.alias_data_dir)
                           for file in files
                           if "_cpd.xml" in file]
        for structure_file in sorted(structure_files):
            self.current_stucture_file = structure_file
            parsed_structure_file = ET.parse(structure_file)
            root_elem = parsed_structure_file.getroot()
            self.make_pointer_new_monograph_title_dict(root_elem)

    def make_pointer_new_monograph_title_dict(self, root_elem):
        if root_elem.find('type').text != "Monograph":
            # Only Monograph types should continue, others exit out now.
            return
        assert self.children_meet_expectations(root_elem,
                                               expected_elems=('type', 'node'),
                                               silent_elems=('node', ))
        node_elems = [child for child in root_elem.iterchildren() if child.tag == 'node']
        for node_elem in node_elems:
            self.loop_one_layer(node_elem)

    def loop_one_layer(self, elem):
        child_node_elems = [child for child in elem.iterchildren() if child.tag == 'node']
        child_page_elems = [child for child in elem.iterchildren() if child.tag == 'page']
        if child_node_elems and not child_page_elems:
            assert self.children_meet_expectations(elem,
                                                   expected_elems=('nodetitle', 'node', 'page'),
                                                   silent_elems=('node', ),)
            for child in child_node_elems:
                self.loop_one_layer(child)
        elif child_page_elems and not child_node_elems:
            self.page_node_bunch(elem)
        else:
            raise Exception('Error:  a page and node on the same level {}'.format(self.current_stucture_file))

    def page_node_bunch(self, node_elem):
        elem_nodetitle = self.get_this_level_nodetitle(node_elem)
        page_elems = [elem for elem in node_elem.iterchildren() if elem.tag == 'page']
        for page_elem in page_elems:
            assert self.children_meet_expectations(page_elem,
                                                   expected_elems=('pageptr', 'pagefile', 'pagetitle'),
                                                   unique_elems=('pageptr', 'pagefile', 'pagetitle'),)
            pointer = page_elem.find('pageptr').text
            title = page_elem.find('pagetitle').text
            if elem_nodetitle:
                new_title = '{} - {}'.format(elem_nodetitle, title)
            else:
                new_title = title
            self.monograph_pointer_newtitle[pointer] = new_title

    def get_this_level_nodetitle(self, elem):
        nodetitle_elem = elem.find('nodetitle')
        if nodetitle_elem is not None:
            return nodetitle_elem.text
        raise Exception('No nodetitle at this level {} {}'.format(elem.tag, self.current_stucture_file))

    def children_meet_expectations(self, elem, expected_elems=[], silent_elems=[], unique_elems=[]):
        unique_set = set()
        for child in elem.iterchildren():
            if child.tag not in expected_elems:
                raise Exception('Unexpected node tag {} {}'.format(child.tag, self.current_stucture_file))
            if child.tag in silent_elems and self.has_text(child):
                raise Exception('Not capturing data in {} {} {}'.format(child.tag, child.text, self.current_stucture_file))
            if child.tag in unique_elems and child.tag in unique_set:
                raise Exception('Duplicate tag {} {}'.format(child.tag, self.current_stucture_file))
            else:
                unique_set.add(child.tag)
        return True

    @staticmethod
    def has_text(elem):
        return elem.text and isinstance(elem.text, str) and len(elem.text.strip()) > 0
